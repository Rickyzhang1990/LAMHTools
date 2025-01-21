from tkinter import *
from tkinter import filedialog, font
import time
from os import path
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color
import tkinter.messagebox

class MyGui:
    def __init__(self,init_window_name):
        self.init_window_name = init_window_name
        self.path = StringVar()
        self.newpath = StringVar()

    def selectPath(self):
        file_path = filedialog.askopenfilename()
        if file_path:
            self.path.set(file_path)

    def selectPath2(self):
        new_path = filedialog.askopenfilename()
        if new_path:
            self.newpath.set(new_path)


    def error(self, errmessage): #错误弹窗
        tkinter.messagebox.showinfo(title= "提示", message= errmessage)

    def uptime(self): #更新时间
        TimeLabel["text"] = datetime.datetime.now().strftime( '%Y-%m-%d %H:%M:%S')
        self.init_window_name.after(1000, self.uptime)
#设置窗口
    def set_init_window(self):
        font1 = font.Font(size=12)
        font2 = font.Font(size=16,weight="bold")
        font3 = font.Font(size=12,weight='bold', )
        self.init_window_name.title("LAMH Tools") #窗口标题栏
        if path.exists('莱盟健康LOGO.ico'):
            self.init_window_name.iconbitmap('莱盟健康LOGO.ico')
        else:
            self.init_window_name.iconbitmap('backup.ico')
        self.init_window_name.geometry( '500x400+400+200') #窗口大小，300x300是窗口的大小，+600是距离左边距的距离，+300是距离上边距的距离
        self.init_window_name.resizable(width=FALSE, height=FALSE) #限制窗口大小，拒绝用户调整边框大小
        Label(self.init_window_name, text= "临床信息更新工具",bg="AliceBlue",fg = "Blue",font=font2).place(relx= 0.31,rely= 0.15) #标签组件，用来显示内容，place里的x、y是标签组件放置的位置
        Label(self.init_window_name, text= "研发临床信息",bg= "AliceBlue",font=font1).place(relx= 0.02,rely= 0.36) #标签组件，用来显示内容，place里的x、y是标签组件放置的位置
        Entry(self.init_window_name, textvariable=self.path).place(relwidth = 0.55,relheight = 0.075,relx= 0.23, rely= 0.35) #输入控件，用于显示简单的文本内容
        Button(self.init_window_name, text= "路径选择", command=self.selectPath,bg= "AliceBlue",font=font1).place(relx= 0.8,rely= 0.35,relwidth=0.16) #按钮组件，用来触发某个功能或函数
        Label(self.init_window_name, text="医学临床信息", bg="AliceBlue", font=font1).place(relx=0.02, rely=0.50)  # 标签组件，用来显示内容，place里的x、y是标签组件放置的位置
        Entry(self.init_window_name, textvariable=self.newpath).place(relwidth=0.55, relheight=0.075, relx=0.23,rely=0.49)  # 输入控件，用于显示简单的文本内容
        Button(self.init_window_name, text="路径选择", command=self.selectPath2, bg="AliceBlue", font=font1).place(
            relx= 0.8, rely= 0.50,relwidth=0.16)
        Button(self.init_window_name, text= "更新表格",bg= "AliceBlue",font=font1,command=self.updateold).place(relwidth= 0.48,relheight= 0.1,relx= 0.27,rely= 0.65) #按钮组件，用来触发某个功能或函数
        self.init_window_name[ "bg"] = "AliceBlue"#窗口背景
        self.init_window_name.attributes( "-alpha", 0.8) #虚化，值越小虚化程度越高


        global TimeLabel #全局变量
        TimeLabel = Label(text="%s"% (datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),bg= "AliceBlue",font=font3) #标签组件，显示时间
        TimeLabel.place(x= 170, y= 330)
        self.init_window_name.after(1000, self.uptime)

        Label(self.init_window_name, text="©Copyright LAMH", fg="red", bg="AliceBlue",
              font=font.Font(size=8, weight="normal", underline=True, slant="italic")).place(relx=0.4, rely=0.95)
#获取当前时间
    def get_current_time(self):
        current_time = time.strftime( '%Y-%m-%d %H:%M:%S',time.localtime(time.time))
        return current_time

    def updateold(self):
        """
        :param sheetname: sheet name of excel
        :return:
        """
        oldname = str(self.path.get())
        date3 = datetime.datetime.now().strftime('%Y-%m-%d')
        names = path.basename(oldname)
        outpath = path.dirname(oldname)
        outname = names.rstrip(".xlsx") if names.endswith("xlsx") else names.rstrip(".xls")
        outname = outpath+"/" + outname + "_" + date3 + ".xlsx"
        old_book = load_workbook(oldname)
        new_book = load_workbook(self.newpath.get())
        sheetname = list(set(old_book.sheetnames) & set(new_book.sheetnames))
        change_font = Font(name='Arial', size=12, bold=True )
        change_fill = PatternFill(start_color=Color(rgb='FFFF0000'), end_color=Color(rgb='FFFF0000'), fill_type='solid')
        change_aliment = Alignment(horizontal='left', vertical='center')
        for name in sheetname:
#            print(name)
            old_sheet = old_book[name]
            new_sheet = new_book[name]
        # 获取旧表格的所有数据（假设数据从第一行开始）
            old_data = {}
            for row in old_sheet.iter_rows(min_row=2, max_row=old_sheet.max_row, min_col=1, max_col=old_sheet.max_column):
                key = row[0].value  # 假设第一列是唯一的键（例如ID）
                old_data[key] = [cell for cell in row]

            # 将新文件中的内容复制到旧文件中，并根据新表格的内容更新旧文件
            for row in new_sheet.iter_rows(min_row=2, max_row=new_sheet.max_row, min_col=1, max_col=new_sheet.max_column):
                key = row[0].value  # 假设第一列是唯一的键（例如ID）

                # 检查key是否在旧数据中
                if key in old_data:
                    # 检查新旧表格内容是否一致
 #                   update_needed = False
 #                   print(len(old_data[key]))
 #                   print(row)
                    for col_idx, cell in enumerate(row[1:], start=2): # 假设从第二列开始是数据
                        old_value = str(old_data[key][col_idx - 1].value).strip()
                        new_value = str(cell.value).strip()
                        if old_value != new_value:  # 比较新旧值
                            old_data[key][col_idx - 1].value = cell.value
                            old_data[key][col_idx - 1].fill  = change_fill
                            old_data[key][col_idx - 1].alignment = change_aliment
                            old_data[key][col_idx - 1].font  = change_font
                        else:
                            pass
                else:
                    # 如果旧表格中没有该条数据，直接将新数据添加到旧表格的最后一行
                    old_sheet.append([cell.value for cell in row])
        if path.exists(outname):
            self.error("已有同名文件，请先删除该文件")
        else:
            old_book.save(outname)
            self.init_window_name.attributes('-topmost', False)
            tkinter.messagebox.showinfo(title="提示", message="结果已生成")
            self.init_window_name.attributes('-topmost', True)



if __name__ == '__main__':
    init_window = Tk() # 实例化出一个父窗口
    LAMH_PORTAL = MyGui(init_window)
    LAMH_PORTAL.set_init_window()
    init_window.mainloop()
